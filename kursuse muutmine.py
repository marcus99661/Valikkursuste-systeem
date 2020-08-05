#### Tegeleb Discordist tulnud käskudega

# 1. Kursusest ära ütlemine, kui mingi muu kursus vajab eemaldatud kursust, siis see tuleb ka ära eemaltada
# 2. Kursuse lisamine (kontroll kas on kohti vaba, õpilasel ei ole juba midagi võetud, eeldusained on jne)
# 3. Saadud kursuste kinnitamine
# 4. ???Doomino effektiga täita tühjad kohad???
# 5. Logi faili loomine muudatustest
'''
class Õpilane:
    def __init__(self, nimi, kursus)
'''
tähed = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
#from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl

def kursuseEemaldamine(nimi, kursus):
    wb = load_workbook("õpilasteFail.xlsx")
    leht = wb['Sheet']
    for i in range(1, 100):
        #print(leht['A' + str(i)].value)
        if leht['A' + str(i)].value == nimi:
            #print("Õpilane nimega: " + nimi + " leitud")
            rida = i
            for j in tähed:
                print(leht[j + str(rida)].value)
                if leht[j + str(rida)].value == kursus:
                    print("Leitud kursus nimega: " + kursus)
                    sheet = wb.get_sheet_by_name("Sheet")
                    sheet[j + str(rida)] = "---"
                    wb.save('õpilasteFail.xlsx')
                    print("Lõpetatud õpilaselt kursuse eemaldamine")
                    break
                elif j == tähed[-1]:
                    print("Ei leitud kursust nimega: " + kursus + ", lõpetan tegevuse")
            break

def kursuseLisamine(nimi, kursus): ## võtab ained.txt kursuse kohad, võtab õpetaja failist õpilaste arvu ja lisab õpilase juurde kui on ruumi, kontrollib eeldusaineid (ja klassi)
    b = load_workbook("õpilasteFail.xlsx")
    leht = wb['Sheet']
    for i in range(1, 100):
        #print(leht['A' + str(i)].value)
        if leht['A' + str(i)].value == nimi:
            #print("Õpilane nimega: " + nimi + " leitud")
            rida = i



#kursuseEemaldamine("Emily 445", "Fotograafia 1")
#kursuseEemaldamine("Emily 445", "Fot")
