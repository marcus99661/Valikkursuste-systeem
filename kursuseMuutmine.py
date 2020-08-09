#### Tegeleb Discordist tulnud käskudega

# 1. Kursusest ära ütlemine, kui mingi muu kursus vajab eemaldatud kursust, siis see tuleb ka ära eemaltada
# 2. Kursuse lisamine (kontroll kas on kohti vaba, õpilasel ei ole juba midagi võetud, eeldusained on jne)
# 3. Saadud kursuste kinnitamine
# 4. ???Doomino effektiga täita tühjad kohad???
# 5. Logi faili loomine muudatustest
'''
class Õpilane:
    def __init__(self, nimi, kursus)
'''
tähed = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]
#from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl

def kursuseEemaldamine(nimi, kursus): ### Peab muutma õpetajateFailis ka, kui võrdleb nimesid ja kursuseid siis võtab ära suure tähe ja tühikud
    õpilaste = load_workbook("õpilasteFail.xlsx")
    leht = õpilaste['Sheet']
    for rida in range(1, leht.max_row):
        #print(leht['A' + str(i)].value)
        if leht['A' + str(rida)].value == nimi:
            #print("Õpilane nimega: " + nimi + " leitud")
            for j in tähed:
                #print(leht[j + str(rida)].value)
                if leht[j + str(rida)].value == kursus:
                    print("Leitud kursus nimega: " + kursus)
                    sheet = õpilaste.get_sheet_by_name("Sheet")
                    sheet[j + str(rida)] = "---"
                    õpilaste.save('õpilasteFail.xlsx')
                    break
                elif j == tähed[-1]:
                    return "Ei leitud kursust nimega: " + kursus + ", lõpetan tegevuse"
                    #print("Ei leitud kursust nimega: " + kursus + ", lõpetan tegevuse")
        elif leht.max_row == rida+1:
            #print(leht.max_row)
            return "Ei leitud õpilast nimega: " + nimi + ", lõpetan tegevuse"
    õpilaste.close()
    ###############
    õpetajate = load_workbook("õpetajateFail.xlsx")
    leht = õpetajate['Sheet']
    for rida in range(1, leht.max_row):
        if leht['A' + str(rida)].value == kursus:
            #print(leht['A' + str(rida)].value)
            #print(leht['E' + str(rida)].value)
            õpilased = leht['E' + str(rida)].value.split(", ")
            try:
                õpilased.remove(nimi)
            except:
                return "ERROR, õpilast ei leitud õpetajate failis kursuse all kirjas"
                #print("ERROR, õpilast ei leitud õpetajate failis kursuse all kirjas")
            seperator = ", "
            sheet = õpilaste.get_sheet_by_name("Sheet")
            sheet['E' + str(rida)] = seperator.join(õpilased)
            print(seperator.join(õpilased))
            õpetajate.save('õpetajateFail.xlsx')

    return 'Edukalt eemaldatud ' + nimi + ' kursuselt "' + kursus + '"'
    #print("Lõpetatud õpilaselt kursuse eemaldamine")
## VAJA LISADA KONTROLLID KAS ÕPILASEL ON EELDUSAINED JNE OLEMAS, ET KURSUST VÕTTA
## 1. võtab kursuse max_õpilaste arvu
## 2. vaatab palju õpilasi on (õpilasedKursusel)
## 3. kontrollib et õpilasel ei oleks sellel perioodil kursust juba võetud
## 4. täidab õpilasel perioodi ära
## 5. lisab õpetajateFaili kursusele õpilase juurde
def kursuseLisamine(nimi, kursus):
    maxÕpilasi = -1
    with open("ained.txt") as file:
        for i in file:
            if i.replace("\n", "").split(";")[0] == kursus:
                maxÕpilasi = int(i.replace("\n", "").split(";")[2]) ## 1.
                periood = int(i.replace("\n", "").split(";")[3])
                if periood == 1:
                    tulp = "B"
                elif periood == 2:
                    tulp = "C"
                elif periood == 3:
                    tulp = "D"
                elif periood == 4:
                    tulp = "E"
                elif periood == 5:
                    tulp = "F"
                elif periood == 6:
                    tulp = "G"
                elif periood == 7:
                    tulp = "H"
                elif periood == 8:
                    tulp = "I"
                break
    if maxÕpilasi == -1:
        return 'Ei leidnud "' + kursus + '" kursuste hulgast'
    
    õpetajate = load_workbook("õpetajateFail.xlsx")
    leht = õpetajate['Sheet']
    for rida in range(1, leht.max_row):
        if leht['A' + str(rida)].value == kursus:
            õpilased = leht['E' + str(rida)].value.split(", ")
            õpilasedKursusel = len(õpilased) ## 2.
            #print(õpilasedKursusel)
            õpetajate.close()
            break
        
    if maxÕpilasi <= õpilasedKursusel:
        return 'Kursus "' + kursus + '" on juba täis ja ei õnnestunud õpilast lisada'
    ###############
    õpilaste = load_workbook("õpilasteFail.xlsx")
    leht = õpilaste['Sheet']
    for rida in range(1, leht.max_row):
        #print(leht['A' + str(i)].value)
        if leht['A' + str(rida)].value == nimi:
            if leht[tulp + str(rida)].value == " --- ": ## 3.
                sheet = õpilaste.get_sheet_by_name("Sheet")
                sheet[tulp + str(rida)] = kursus ## 4. 
                õpilaste.save('õpilasteFail.xlsx') ################### SEEE KOHT ÕLE KONTROLLIDA
            else:
                return 'Õpilasel ' + nimi + ' on juba sellel perioodil võetud "' + leht[tulp + str(rida)].value + '"' 
        elif leht.max_row == rida+1:
            return 'Õpilast ' + nimi + ' ei leitud õpilaste nimekirjast' 
    ###############
    õpetajate = load_workbook("õpetajateFail.xlsx")
    leht = õpetajate['Sheet']
    for rida in range(1, leht.max_row):
        if leht['A' + str(rida)].value == kursus:
            #print(leht['A' + str(rida)].value)
            #print(leht['E' + str(rida)].value)
            õpilased = leht['E' + str(rida)].value.split(", ")
            try:
                õpilased.append(nimi)
            except:
                return "ERROR, õpilast ei leitud õpetajate failis kursuse all kirjas"
                #print("ERROR, õpilast ei leitud õpetajate failis kursuse all kirjas")
            seperator = ", "
            sheet = õpilaste.get_sheet_by_name("Sheet")
            sheet['E' + str(rida)] = seperator.join(õpilased)
            #print(seperator.join(õpilased))
            õpetajate.save('õpetajateFail.xlsx')
            return 'Edukalt lisatud ' + nimi + ' kursusele "' + kursus + '"'


    '''
    õpilaste = load_workbook("õpilasteFail.xlsx")
    leht = õpilaste['Sheet']
    for i in range(1, 100):
        #print(leht['A' + str(i)].value)
        if leht['A' + str(i)].value == nimi:
            #print("Õpilane nimega: " + nimi + " leitud")
            rida = i
    '''

if __name__ == "__main__":

    #print(kursuseLisamine("asdasdasd", "Fotograafia 1"))
    print(kursuseLisamine("Isabella 474", "Programmeerimine keeles Python 2"))
    print(kursuseLisamine("asdasdasd", "Fotogasdasdraafia 1"))
    #print(kursuseEemaldamine("asdasdasd", "Fotograafia 1"))
    #kursuseEemaldamine("Emily 445", "Fotograafia 1")
    #kursuseEemaldamine("Emily 445", "Fot")
    pass
