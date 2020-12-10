#        VAJA PARANDADA
###################################
# OSAD KURSUSED EI OLE KÕIGILE KLASSIDELE
# EELMINE AASTA ON VÕETUD EELDUSAINE
# VAJA LISADA PRAKTIKUMID/TUNNIVÄLISED (ühe valiku all, valid nii palju kui tahad)
# 5 KURSUST IGALE ÕPILASELE MIINIMUM
# ÕPILASTEL KURSUSTE ÜMBERTÕSTMINE
# VAJA LUUA ÕPILASTELE OUTPUT FILE  -  TEHTUD
# VAJA LUUA ÕPETAJATELE OUTPUT FILE  -  TEHTUD
# LISADA KOMMENTAARE
# LUUA DOKUMENT
# LOGI FAILI LOOMINE - TEHTUD
###################################

import random
from datetime import datetime
from csv import reader
from collections import OrderedDict
import copy
from openpyxl import Workbook

#{'Marcus': {'Ajatempel': '2020/07/04 10:41:27 PM GMT +3', 'Kasutajanimi': 'marcus99661@gmail.com', 'Nimi': 'Marcus', 'Klass': '10', '2. periood hommik 1. valik': 'Aine 1', '2. periood hommik 2. valik': 'Aine 2', '2. periood hommik 3. valik': 'Ei taha', '2. periood Ãµhtu 1. valik': 'Ei taha', 'EI VASTA': ''}}

def kirjuta(tekst):
    print(datetime.now().strftime("[%H:%M:%S.%f]") + " " + tekst)
    file.write(datetime.now().strftime("[%H:%M:%S.%f]") + " " + tekst + "\n")

def grupeerimine():
    klass10 = {}
    klass11 = {}
    klass12 = {}
    formating = []
    õpilaseNimed = []
    temp1 = {}
    korda = 0
    with open('testinputpraks.csv', 'r', encoding="utf-8") as input_file: # Õpilaste kursuste faili avamine
        csv_reader = reader(input_file)
        for row in csv_reader:
            #print(row)
            if korda == 0:
                formating = row
                for i in range(len(formating)):
                    formating[i] = formating[i].replace('"', "")
                korda += 1
            else:
                for i in range(len(row)):
                    row[i] = row[i].replace('"', "") # Encodingu parandus
                    row[i] = row[i].replace('Ãµ', "õ")
                    row[i] = row[i].replace('Ã¤', "ä")
                    row[i] = row[i].replace('Ã¶', "ö")
                    row[i] = row[i].replace('Ã¼', "ü")
                
                #### Paneb õige klassi sõnastikku
                õpilaseNimed.append(row[2]) # Sorteerib õpilased õigetesse klassidesse
                if row[3] == '12':
                    for i in range(len(formating)):
                        temp1[formating[i]] = row[i]
                    klass12[row[2]] = temp1
                elif row[3] == '11':
                    for i in range(len(formating)):
                        temp1[formating[i]] = row[i]
                    klass11[row[2]] = temp1
                elif row[3] == '10':
                    for i in range(len(formating)):
                        temp1[formating[i]] = row[i]
                    klass10[row[2]] = temp1
                else:
                    kirjuta("TEKKIS VIGA ÕPILASE ÕIGESSE SÕNASTIKKU PANEMISEL")
                temp1 = {}
    return klass10, klass11, klass12, õpilaseNimed

klass10, klass11, klass12, õpilaseNimed = grupeerimine() # "õpilaseNimed" ei ole hetkel vajalik/kasutuses
#print(klass10)


#klass10_PERM()
ained = {}
def ained_seadistamine(): # Kursuste seadistamine kergeks kasutuseks
    ainedList = []
    ained = {}
    with open('kursused.txt', 'r', encoding="utf-8") as ainedfail:
        for i in ainedfail: #### {Python 2 : {"kohad" : 40, "eeldusaine" = "Python 1", "lisad" : ""}}
            temp1 = {}
            ainedList = i.replace("\n", "").split(";")
            #ained[ainedList[0]]["kohad"] = ainedList[1]
            #ained[ainedList[0]]["kohad"] = 40
            temp1["alternatiiv"] = ainedList[1]
            temp1["kohad"] = ainedList[2]
            temp1["periood"] = ainedList[3]
            temp1["kohtiVõetud"] = 0
            temp1["eeldusaine"] = ainedList[4]
            temp1["üksEeldusaine"] = ainedList[5]
            temp1["lisad"] = ainedList[6]
            temp1["vastuVõetud"] = []
            temp1["järjekord"] = []
            ained[ainedList[0]] = temp1
            #print(ained)
        #print(ained)
        return ained


klass10_seadistatud, klass11_seadistatud, klass12_seadistatud = {}, {}, {}


#### 12.klassi 1. valik kirja 

def seadistamine(õpilaseNimi, õpilaneDict):
    õpilaseKursused = [] #### [[P1H väga, P1H võtaks, P1H vähe], [P1Õ väga, P1õ võtaks, P1Õ vähe]]
    ajutine = []
    korda1 = 0
    #print(õpilaseNimi)
    for i in range(4, len(list(õpilaneDict))): ######################### KUI TULEB EMAIL KONTROLL VB PEAB LISAMA -1 len() LÕPPU
        if korda1 % 3 == 0 and korda1 != 0:
            õpilaseKursused.append(ajutine)
            ajutine = []
        korda1 += 1
        ajutine.append(list(õpilaneDict.values())[i])
    õpilaseKursused.append(ajutine)
    #print(õpilaseKursused)
    return õpilaseKursused


for i in range(0, len(klass12)): # Õpilaste sidumine nende valitud kursustega
    õpilaseNimi, õpilaseDict = random.choice(list(klass12.items()))
    del klass12[õpilaseNimi]
    klass12_seadistatud[õpilaseNimi] = seadistamine(õpilaseNimi, õpilaseDict)

#print(klass12_seadistatud)

for i in range(0, len(klass11)):
    õpilaseNimi, õpilaseDict = random.choice(list(klass11.items()))
    del klass11[õpilaseNimi]
    klass11_seadistatud[õpilaseNimi] = seadistamine(õpilaseNimi, õpilaseDict)

for i in range(0, len(klass10)):
    õpilaseNimi, õpilaseDict = random.choice(list(klass10.items()))
    del klass10[õpilaseNimi]
    klass10_seadistatud[õpilaseNimi] = seadistamine(õpilaseNimi, õpilaseDict)

klass10_PERM, klass11_PERM, klass12_PERM = copy.deepcopy(klass10_seadistatud), copy.deepcopy(klass11_seadistatud), copy.deepcopy(klass12_seadistatud)
ained = ained_seadistamine()
#print(ained)

resultList = {} ####{Marcus : ['Planimeetria alused', '3D-modelleerimine']}
result = {} ####{Marcus : {1 : 'Planimeetria alused', 2 : '', 3 : '3D-modelleerimine'}}
tempSõnastik = {}

def võrdlemine(a, b): 
    for i in a:
        if i in b:
            return True
    return False

#### [['Planimeetria alused', 'Globaliseeruv maailm', 'Ei taha'], ['Programmeerimine keeles Python 1', 'Programmeerimine keeles Python 1', 'Programmeerimine keeles Python 1'], ['3D-modelleerimine', 'Küberkaitse 1', '3D-modelleerimine'], ['Statistiline maailmapilt', 'Ei taha', 'Ei taha'], ['Loomade käitumine', 'Ei taha', 'Ei taha'], ['Majandusõpe', 'Millest ELU koosneb?', 'Majandusõpe'], ['Ettevõtlusõpe', 'Liiklusfüüsika', 'Liiklusfüüsika'], ['CAD joonestamine', 'Ei taha', 'Ei taha']]
with open("log.txt", "w") as file:
    file.write("")
    def registreerimine(klass_seadistatud, valik, result):
        for i in range(0, len(klass_seadistatud)):
            temp2 = []
            tempSõnastik = {}
            klass_seadistatud_temp = klass_seadistatud
            õpilaseNimi, õpilaseKursused = random.choice(list(klass_seadistatud_temp.items()))
            del klass_seadistatud_temp[õpilaseNimi]
            if õpilaseNimi not in result:
                tempSõnastik = {k: '' for k in range(1, len(õpilaseKursused))} ## tempSõnastik = {k: '' for k in range(1, len(õpilaseKursused)+1)}
                result[õpilaseNimi] = tempSõnastik
            for i in range(0, len(õpilaseKursused)-1): ### eemaldada -1
                #print(õpilaseKursused)
                hetkeneKursus = õpilaseKursused[i][valik]
                temp2 = resultList.get(õpilaseNimi, [])
                resultList[õpilaseNimi] = temp2
                if hetkeneKursus == "Ei taha":
                    kirjuta(õpilaseNimi + " ei tahtnud see periood midagi")
                elif ained[hetkeneKursus]['kohtiVõetud'] <  int(ained[hetkeneKursus]['kohad']):  # 1. kui mahub 2. ei ole juba sellel kursusel 3. ei ole see periood veel midagi võetud 4. eeldusained on võetud (eelmine periood või eelmine aasta) 5. üks sama eeldusaine on võetud
                    if hetkeneKursus not in resultList[õpilaseNimi]: #### õpilasel on see kursus juba võetud
                        if ained[hetkeneKursus]['alternatiiv'] not in resultList[õpilaseNimi]: #### õpilasel on selle kursuse alternatiiv võetud
                            if result[õpilaseNimi][int(ained[hetkeneKursus]['periood'])] == '': #### õpilasel on sellel perioodil midagi juba võetud
                                #### TULEB LISADA EELMISE AASTA EELDUSAINED JUURDE
                                if ained[hetkeneKursus]['eeldusaine'] == '' or all(elem in resultList[õpilaseNimi] for elem in ained[hetkeneKursus]['eeldusaine'].split(",")): ##### kontrollib kas õpilane on see aasta võtnud eeldusained
                                    if ained[hetkeneKursus]['üksEeldusaine'] == '' or võrdlemine(ained[hetkeneKursus]['üksEeldusaine'].split(","), resultList[õpilaseNimi]):
                                        kirjuta(õpilaseNimi + " registreeritud kursusele " + hetkeneKursus)
                                        ained[hetkeneKursus]['kohtiVõetud'] += 1
                                        ained[hetkeneKursus]['vastuVõetud'].append(õpilaseNimi)
                                        ########
                                        temp2 = resultList.get(õpilaseNimi, [])
                                        temp2.append(hetkeneKursus)
                                        resultList[õpilaseNimi] = temp2
                                        #print(resultList)
                                        ########
                                        tempSõnastik = result[õpilaseNimi]
                                        tempSõnastik[i+1] = hetkeneKursus
                                        result[õpilaseNimi] = tempSõnastik
                                        #print(result)
                                        ########
                                        if ained[hetkeneKursus]['lisad'] != '':
                                            for j in range(0, len(ained[hetkeneKursus]['lisad'].split(","))):
                                                kirjuta(õpilaseNimi + " vastu võetud " + ained[hetkeneKursus]['lisad'].split(",")[j])
                                                ained[ained[hetkeneKursus]['lisad'].split(",")[j]]['vastuVõetud'].append(õpilaseNimi) ## Kui õigesti mäletan siis see rida lisab kursuse "vastuVõetud" listi õpilase nime juurde
                                                #####
                                                temp2 = resultList.get(õpilaseNimi, [])
                                                temp2.append(ained[hetkeneKursus]['lisad'].split(",")[j])
                                                resultList[õpilaseNimi] = temp2
                                                #print(resultList)
                                                #####
                                                tempSõnastik[int(ained[ained[hetkeneKursus]['lisad'].split(",")[j]]['periood'])] = ained[hetkeneKursus]['lisad'].split(",")[j]
                                                result[õpilaseNimi] = tempSõnastik
                                                #print(result)
                                                #####
                                else:
                                    kirjuta(õpilaseNimi + ' ei saanud "' + hetkeneKursus + '", sest ei ole võtnud eeldusainet ' + ained[hetkeneKursus]['eeldusaine'])
                            else:
                                kirjuta(õpilaseNimi + ' ei saanud kursusele "' + hetkeneKursus + '", sest on juba sellel perioodil muule kursusele sisse saanud')
                        else:
                            kirjuta(õpilaseNimi + ' on juba "' + hetkeneKursus + '" kursuse alternatiivile sisse eelneval hetkel')
                    else:
                        kirjuta(õpilaseNimi + ' on juba "' + hetkeneKursus + '" kursusele sisse saanud eelneval hetkel')
                else:
                    kirjuta(õpilaseNimi + ' ei saanud kursusele "' + hetkeneKursus + '", sest see oli juba täis')
                    if len(ained[hetkeneKursus]["järjekord"]) < 10:
                        ained[hetkeneKursus]["järjekord"].append(õpilaseNimi)
                        kirjuta(õpilaseNimi + ' lisati "' + hetkeneKursus + '" järjekorda')

        #print(result)
        return result
    #### 12. klassi 1. valik valimine
    klass12_segamini = {}
    for i in range(0, len(klass12_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass12_seadistatud))
        kursused = klass12_seadistatud[õpilaseNimi]
        del klass12_seadistatud[õpilaseNimi]
        klass12_segamini[õpilaseNimi] = kursused
    result = registreerimine(klass12_segamini, 0,result)
    print("LÕPETATUD 12. KLASSI 1. VALIK")
    file.write("LÕPETATUD 12. KLASSI 1. VALIK")
    #### 11. ja 10. klassi 1. valik
    klass12_seadistatud = copy.deepcopy(klass12_PERM)
    kokku11ja10 = {} #### PANEB KÕIK 11 JA 10 KLASSI ÕPILASED ÜHTE SÕNASTIKKU SUVALISES JÄRJEKORRAS JA HAKKAB REGISTREERIMA
    kokku11ja10_segamini = {}

    for i in range(0, len(klass11_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass11_seadistatud))
        kursused = klass11_seadistatud[õpilaseNimi]
        del klass11_seadistatud[õpilaseNimi]
        kokku11ja10[õpilaseNimi] = kursused
    for i in range(0, len(klass10_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass10_seadistatud))
        kursused = klass10_seadistatud[õpilaseNimi]
        del klass10_seadistatud[õpilaseNimi]
        kokku11ja10[õpilaseNimi] = kursused

    for i in range(0, len(kokku11ja10.keys())):
        õpilaseNimi = random.choice(list(kokku11ja10))
        kursused = kokku11ja10[õpilaseNimi]
        del kokku11ja10[õpilaseNimi]
        kokku11ja10_segamini[õpilaseNimi] = kursused
    result = registreerimine(kokku11ja10_segamini, 0,result)
    print("LÕPETATUD 11. JA 10. KLASSI 1. VALIK")
    file.write("LÕPETATUD 11. JA 10. KLASSI 1. VALIK")
    #### 12., 11. ja 10. klassi 2. valik
    klass12_seadistatud = copy.deepcopy(klass12_PERM)
    klass11_seadistatud = copy.deepcopy(klass11_PERM)
    klass10_seadistatud = copy.deepcopy(klass10_PERM)
    kokku12ja11ja10 = {}
    kokku12ja11ja10_segamini = {}

    for i in range(0, len(klass12_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass12_seadistatud))
        kursused = klass12_seadistatud[õpilaseNimi]
        del klass12_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
    for i in range(0, len(klass11_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass11_seadistatud))
        kursused = klass11_seadistatud[õpilaseNimi]
        del klass11_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
    for i in range(0, len(klass10_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass10_seadistatud))
        kursused = klass10_seadistatud[õpilaseNimi]
        del klass10_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
        
    for i in range(0, len(kokku12ja11ja10.keys())):
        õpilaseNimi = random.choice(list(kokku12ja11ja10))
        kursused = kokku12ja11ja10[õpilaseNimi]
        del kokku12ja11ja10[õpilaseNimi]
        kokku12ja11ja10_segamini[õpilaseNimi] = kursused

    result = registreerimine(kokku12ja11ja10_segamini, 1,result)
    #print(kokku12ja11ja10_segamini)
    print("LÕPETATUD 12., 11. ja 10. KLASSI 2. VALIK")
    file.write("LÕPETATUD 12., 11. ja 10. KLASSI 2. VALIK")
    #### 12., 11. ja 10. klassi 3. valik
    klass12_seadistatud = copy.deepcopy(klass12_PERM)
    klass11_seadistatud = copy.deepcopy(klass11_PERM)
    klass10_seadistatud = copy.deepcopy(klass10_PERM)

    kokku12ja11ja10 = {}
    kokku12ja11ja10_segamini = {}
    for i in range(0, len(klass12_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass12_seadistatud))
        kursused = klass12_seadistatud[õpilaseNimi]
        del klass12_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
    for i in range(0, len(klass11_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass11_seadistatud))
        kursused = klass11_seadistatud[õpilaseNimi]
        del klass11_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
    for i in range(0, len(klass10_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass10_seadistatud))
        kursused = klass10_seadistatud[õpilaseNimi]
        del klass10_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
        
    for i in range(0, len(kokku12ja11ja10.keys())):
        õpilaseNimi = random.choice(list(kokku12ja11ja10))
        kursused = kokku12ja11ja10[õpilaseNimi]
        del kokku12ja11ja10[õpilaseNimi]
        kokku12ja11ja10_segamini[õpilaseNimi] = kursused
        
    result = registreerimine(kokku12ja11ja10_segamini, 2,result)
    #print(kokku12ja11ja10_segamini)
    print("LÕPETATUD 12., 11. ja 10. KLASSI 3. VALIK")
    file.write("LÕPETATUD 12., 11. ja 10. KLASSI 3. VALIK")

    print(result)
    print("-------------------------")

    for i in range(0, len(klass10_PERM.keys())):
        õpilaseNimi = list(klass10_PERM)[i]
        kursused = klass10_PERM[õpilaseNimi]
        print("Õpilasel " + õpilaseNimi + " on võetud praktikumid: " + ', '.join(''.join(kursused[-1]).split(";")))
        
    klass12_seadistatud = copy.deepcopy(klass12_PERM)
    klass11_seadistatud = copy.deepcopy(klass11_PERM)
    klass10_seadistatud = copy.deepcopy(klass10_PERM)
    kokku12ja11ja10 = {}
    kokku12ja11ja10_segamini = {}

    for i in range(0, len(klass12_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass12_seadistatud))
        kursused = klass12_seadistatud[õpilaseNimi]
        del klass12_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
    for i in range(0, len(klass11_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass11_seadistatud))
        kursused = klass11_seadistatud[õpilaseNimi]
        del klass11_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
    for i in range(0, len(klass10_seadistatud.keys())):
        õpilaseNimi = random.choice(list(klass10_seadistatud))
        kursused = klass10_seadistatud[õpilaseNimi]
        del klass10_seadistatud[õpilaseNimi]
        kokku12ja11ja10[õpilaseNimi] = kursused
        
    for i in range(0, len(kokku12ja11ja10.keys())):
        õpilaseNimi = random.choice(list(kokku12ja11ja10))
        #kursused = kokku12ja11ja10[õpilaseNimi][-1]
        kursused = "".join(kokku12ja11ja10[õpilaseNimi][-1]).split(";")
        del kokku12ja11ja10[õpilaseNimi]
        kokku12ja11ja10_segamini[õpilaseNimi] = kursused

    print(kokku12ja11ja10_segamini)
    #result = registreerimine(kokku12ja11ja10_segamini, 1,result)

    def praks_seadistamine(): # Kursuste seadistamine kergeks kasutuseks
        ainedList = []
        ained = {}
        with open('praks.txt', 'r', encoding="utf-8") as ainedfail:
            for i in ainedfail: #### 
                temp1 = {}
                ainedList = i.replace("\n", "").split(";")
                #ained[ainedList[0]]["kohad"] = ainedList[1]
                #ained[ainedList[0]]["kohad"] = 40
                temp1["kohad"] = ainedList[1]
                temp1["kohtiVõetud"] = 0
                temp1["eeldusaine"] = ainedList[2]
                temp1["üksEeldusaine"] = ainedList[3]
                temp1["samalAjal"] = ainedList[4]
                temp1["sarnased"] = ainedList[5]
                temp1["vastuVõetud"] = []
                temp1["järjekord"] = []
                ained[ainedList[0]] = temp1
                #print(ained)
            #print(ained)
            return ained

    praks = praks_seadistamine()
    print(praks)
    
    def praksSamalAjal(õpilaseNimi, suvalinepraks, praks):
        for i in range(len(praks.keys())):
            asd = praks[list(praks.keys())[i]]
            praksNimi = list(praks.keys())[i]
            if suvalinepraks in asd["samalAjal"] and praksNimi in praksResult[õpilaseNimi]:
                return False
        return True

    ### PRAKS REGISTREERIMINE

    def praks_registreerimine(õpilaseNimi, suvalinePraks, result): ########## Kõik kontrollid ja järjekorda lisamine, uute testandmete loomine
        if praks[suvalinePraks]["kohtiVõetud"] != praks[suvalinePraks]["kohad"]: #### kontrollib kas praks on täis
            if praksSamalAjal(õpilaseNimi, suvalinePraks, praks): #### kontrollib kas praks toimub samal ajal kui õpilasel mingi teine kursus
                pass


        praks[suvalinePraks]["kohtiVõetud"] += 1
        praks[suvalinePraks]["vastuVõetud"].append(õpilaseNimi)
        
        if õpilaseNimi in result.keys():
            result[õpilaseNimi].append(suvalinePraks)
        else:
            result[õpilaseNimi] = []
            result[õpilaseNimi].append(suvalinePraks)
        #result[õpilaseNimi] = suvalinePraks
        return result
    
    
    praksResult = {}   #### {Marcus : ["Praksid", "Praksid 2"]}

    while True:
        if len(kokku12ja11ja10_segamini) <= 0:
            break
        õpilaseNimi = random.choice(list(kokku12ja11ja10_segamini))
        õpilasePraksid = kokku12ja11ja10_segamini[õpilaseNimi] ## eemaldamiseks
        #print("Järgmine on: " + õpilaseNimi + " " + str(õpilasePraksid))
        if õpilasePraksid == [''] or õpilasePraksid == [] or õpilasePraksid[0] == "": ### Õpilaste kellel pole prakse eemaldamine
            del kokku12ja11ja10_segamini[õpilaseNimi]
            continue
        õpilaseSuvalinePraks = random.choice(list(kokku12ja11ja10_segamini[õpilaseNimi]))
        
        praksResult = praks_registreerimine(õpilaseNimi, õpilaseSuvalinePraks, praksResult)
        #print(praksResult)
        õpilasePraksid.remove(õpilaseSuvalinePraks)
        kokku12ja11ja10_segamini[õpilaseNimi] = õpilasePraksid
    
    print(praksResult)
    
    print(praks) ### prakside kohta info

    #################################### ÕPETAJA FAILI KIRJUTAMINE
    kursusedFailiJaoks = []
    praksidFailiJaoks = []
    õpilasteNimedFailiJaoks = []
    praksÕpilasteNimedFailiJaoks = []
    järjekordFailiJaoks = []
    praksJärjekordFailiJaoks = []
    resultKeys = ained.keys()
    for i in range(0, len(resultKeys)):
        kursusedFailiJaoks.append(list(resultKeys)[i])
        õpilasteNimedFailiJaoks.append(ained[list(resultKeys)[i]]["vastuVõetud"])
        järjekordFailiJaoks.append(ained[list(resultKeys)[i]]["järjekord"])
    
    for i in range(0, len(praks.keys())):
        praksidFailiJaoks.append(list(praks.keys())[i])
        praksÕpilasteNimedFailiJaoks.append(praks[list(praks.keys())[i]]["vastuVõetud"])
        praksJärjekordFailiJaoks.append(praks[list(praks.keys())[i]]["järjekord"])

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Kursus"
    sheet["E1"] = "Õpilased"
    sheet["O1"] = "Järjekord"
    seperator = ", "
    for i in range(0, len(kursusedFailiJaoks)):
        sheet.cell(row=i+2, column=1).value = kursusedFailiJaoks[i]
        sheet.cell(row=i+2, column=5).value = seperator.join(õpilasteNimedFailiJaoks[i])
        sheet.cell(row=i+2, column=15).value = seperator.join(järjekordFailiJaoks[i]) ########## LISADA ÕPILASTE JÄRJEKORD

    sheet.cell(row=len(kursusedFailiJaoks)+2, column=1).value = "PRAKTIKUMID:"
    print(praksidFailiJaoks)
    for i in range(len(kursusedFailiJaoks)+1, len(praksidFailiJaoks) + len(kursusedFailiJaoks) - 1):
        sheet.cell(row=i+2, column=1).value = praksidFailiJaoks[i - len(kursusedFailiJaoks)+1]
        sheet.cell(row=i+2, column=5).value = seperator.join(praksÕpilasteNimedFailiJaoks[i - len(kursusedFailiJaoks)+1])
        sheet.cell(row=i+2, column=15).value = seperator.join(praksJärjekordFailiJaoks[i - len(kursusedFailiJaoks)+1])

    workbook.save(filename="õpetajateFail.xlsx")
    print("-------------------------")
    #################################### ÕPILASTE FAILI KIRJUTAMINE

    igaÕpilaseKursused = []
    igaÕpilasePraksid = []
    #õpilaseNimed = []
    resultKeys = result.keys() 
    for i in range(0, len(resultKeys)):
        #igaÕpilaseKursused.append(result[list(resultKeys)[i]])
        hetkesedKursused = []
        for j in range(1, int(list(result[list(resultKeys)[0]].keys())[-1])+1):
            if result[list(resultKeys)[i]].get(int(j)) == "":
                hetkesedKursused.append(" --- ")
            else:
                hetkesedKursused.append(result[list(resultKeys)[i]].get(int(j)))
        igaÕpilaseKursused.append(hetkesedKursused)

    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Õpilane"
    sheet["B1"] = "2. periood hommik"
    sheet["C1"] = "2. periood õhtu"
    sheet["D1"] = "3. periood hommik"
    sheet["E1"] = "3. periood õhtu"
    sheet["F1"] = "4. periood hommik"
    sheet["G1"] = "4. periood õhtu"
    sheet["H1"] = "5. periood hommik"
    sheet["I1"] = "5. periood õhtu"
    sheet["J1"] = "Praktikumid ja tunnivälised kursused"

    seperator = ", "
    #print(igaÕpilaseKursused)
    for i in range(0, len(igaÕpilaseKursused)):
        nimi = list(resultKeys)[i]
        sheet["A" + str(i+2)] = nimi
        sheet["B" + str(i+2)] = igaÕpilaseKursused[i][0]
        sheet["C" + str(i+2)] = igaÕpilaseKursused[i][1]
        sheet["D" + str(i+2)] = igaÕpilaseKursused[i][2]
        sheet["E" + str(i+2)] = igaÕpilaseKursused[i][3]
        sheet["F" + str(i+2)] = igaÕpilaseKursused[i][4]
        sheet["G" + str(i+2)] = igaÕpilaseKursused[i][5]
        sheet["H" + str(i+2)] = igaÕpilaseKursused[i][6]
        sheet["I" + str(i+2)] = igaÕpilaseKursused[i][7]
        sheet["J" + str(i+2)] = seperator.join(praksResult.setdefault(nimi, ""))#praksResult
    workbook.save(filename="õpilasteFail.xlsx")
    print("LÕPETATUD EDUKALT ÕPILASTE KURSUSELE LISAMISE")
    file.write("LÕPETATUD EDUKALT ÕPILASTE KURSUSELE LISAMISE")
    '''
    for i in range(0, len(klass10_PERM.keys())):
        õpilaseNimi = list(klass10_PERM)[i]
        kursused = klass10_PERM[õpilaseNimi]
        print("Õpilasel " + õpilaseNimi + " on võetud praktikumid: " + ', '.join(''.join(kursused[-1]).split(";")))
    '''

    print("LÕPETATUD EDUKALT")
    file.write("LÕPETATUD EDUKALT")
